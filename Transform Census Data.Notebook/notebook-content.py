# Fabric notebook source

# METADATA ********************

# META {
# META   "kernel_info": {
# META     "name": "synapse_pyspark"
# META   },
# META   "dependencies": {
# META     "lakehouse": {
# META       "default_lakehouse": "68cab2d5-d6ec-47a8-a3ce-904a41379bf5",
# META       "default_lakehouse_name": "kipu_lakehouse",
# META       "default_lakehouse_workspace_id": "fb72ebcf-98cc-4162-85c9-5d2042b8b795",
# META       "known_lakehouses": [
# META         {
# META           "id": "68cab2d5-d6ec-47a8-a3ce-904a41379bf5"
# META         }
# META       ]
# META     }
# META   }
# META }

# CELL ********************

# =============================================================================
# KIPU CENSUS DATA PIPELINE — ALL FACILITIES
# =============================================================================
# Reads raw census JSON from blob storage for all 7 facilities,
# applies mappings from Excel file, and writes 4 core Delta tables.
#
# TABLES CREATED:
#   census_raw        — raw flattened patient records (all facilities)
#   census_clean      — deduplicated, mapped, enriched patient-day records
#   budget            — monthly ADC budget targets (Longbranch only for now)
#   data_validation   — unmapped value counts and data quality status
#
# PROGRAM MAPPING:
#   Uses 3-column key: Program + Facility + Location for precise matching.
#   Falls back to Program-only match if no facility+location match found.
#
# INCREMENTAL LOGIC:
#   Checks max census_date per facility in census_raw.
#   Only processes files newer than last processed date per facility.
#   census_clean always fully rebuilt from census_raw.
#
# TO RUN MANUALLY: Run all cells top to bottom.
# SCHEDULED: Daily at 1AM Central via Fabric notebook schedule.
# =============================================================================

from pyspark.sql.functions import *
from pyspark.sql.types import *
from datetime import datetime as dt, timezone, timedelta
from functools import reduce
import pandas as pd
import os

spark.conf.set("spark.databricks.delta.schema.autoMerge.enabled", "true")

# =============================================================================
# FACILITY CONFIG
# =============================================================================

FACILITY_SHORTCUTS = [
    # (local_path, spark_path, facility_name)
    # Longbranch has two shortcuts — historical backfill + live
    ("/lakehouse/default/Files/Longbranch Daily 03.30.26", "Files/Longbranch Daily 03.30.26", "Longbranch"),
    ("/lakehouse/default/Files/Longbranch Daily",          "Files/Longbranch Daily",          "Longbranch"),
    ("/lakehouse/default/Files/RNGA Daily",                "Files/RNGA Daily",                "RNGA"),
    ("/lakehouse/default/Files/Tides Edge Daily",          "Files/Tides Edge Daily",          "Tides Edge Recovery"),
    ("/lakehouse/default/Files/Lotus Daily",               "Files/Lotus Daily",               "Lotus Wellness"),
    ("/lakehouse/default/Files/CRC Daily",                 "Files/CRC Daily",                 "Chattanooga Recovery Center"),
    ("/lakehouse/default/Files/Graceland Daily",           "Files/Graceland Daily",           "Graceland Recovery"),
    ("/lakehouse/default/Files/Green Acres Daily",         "Files/Green Acres Daily",         "Green Acres Recovery"),
    ("/lakehouse/default/Files/RNAL Daily",                "Files/RNAL Daily",                "RNAL"),
]

# =============================================================================
# STEP 1: LOAD MAPPINGS FROM EXCEL
# =============================================================================

MAPPING_FILE = "/lakehouse/default/Files/Mappings/Longbranch_KIPU_Mapping.xlsx"

def load_program_mapping():
    """
    Load program mapping using 3-column key: Program + Facility + Location.
    Rows with blank Facility/Location act as fallback (match any facility/location).
    """
    df = pd.read_excel(MAPPING_FILE, sheet_name="program_mapping", header=1)
    df = df[df["Status"].astype(str).str.startswith("✅")]
    df = df.dropna(subset=["Program Group"])
    df = df[df["Program Group"].astype(str).str.strip() != ""]
    rows = []
    for _, r in df.iterrows():
        program  = str(r["Program Name (from Kipu)"]).strip() if pd.notna(r["Program Name (from Kipu)"]) else "NULL"
        facility = str(r["Facility"]).strip()  if pd.notna(r.get("Facility"))  and str(r.get("Facility")).strip() not in ("", "nan") else ""
        location = str(r["Location"]).strip()  if pd.notna(r.get("Location"))  and str(r.get("Location")).strip() not in ("", "nan") else ""
        group    = str(r["Program Group"]).strip()
        rows.append((program, facility, location, group))
    return rows

def load_mapping(sheet_name, key_col, value_col):
    """Simple single-key mapping for insurance and discharge."""
    df = pd.read_excel(MAPPING_FILE, sheet_name=sheet_name, header=1)
    df = df.dropna(subset=[key_col, value_col])
    df = df[df[value_col].astype(str).str.strip() != ""]
    df = df[~df[value_col].astype(str).str.startswith("⚠️")]
    return dict(zip(df[key_col].astype(str).str.strip(), df[value_col].astype(str).str.strip()))

def load_budget():
    """Read budget tab — melt wide format to long."""
    df = pd.read_excel(MAPPING_FILE, sheet_name="budget", header=1)
    df = df.dropna(subset=["Month"])
    categories = [c for c in df.columns if c not in ["Month", "Total ADC"]]
    rows = []
    for _, row in df.iterrows():
        for cat in categories:
            val = row.get(cat)
            if pd.notna(val):
                rows.append((str(row["Month"]).strip(), cat, float(val)))
    return rows

print("Loading mappings from Excel...")
program_mapping_rows = load_program_mapping()
insurance_mapping    = load_mapping("insurance_mapping", "Insurance Company (from Kipu)", "Payor Group")
discharge_mapping    = load_mapping("discharge_mapping", "Discharge Type (from Kipu)", "Discharge Group")
budget_rows          = load_budget()

print(f"  Program mappings:   {len(program_mapping_rows)}")
print(f"  Insurance mappings: {len(insurance_mapping)}")
print(f"  Discharge mappings: {len(discharge_mapping)}")
print(f"  Budget rows:        {len(budget_rows)}")
print("✅ Mappings loaded")

# =============================================================================
# STEP 2: DETERMINE LAST PROCESSED DATE PER FACILITY (INCREMENTAL)
# =============================================================================

try:
    existing = spark.table("census_raw")
    facility_dates = existing.groupBy("facility").agg(
        max("census_date").alias("last_date")
    ).collect()
    last_date_per_facility = {row["facility"]: row["last_date"] for row in facility_dates}
    incremental = True
    print("Incremental mode. Last processed dates:")
    for fac, d in sorted(last_date_per_facility.items()):
        print(f"  {fac}: {d}")
except Exception:
    last_date_per_facility = {}
    incremental = False
    print("No existing census_raw — full load for all facilities")

def get_process_from(facility_name):
    last = last_date_per_facility.get(facility_name)
    if last:
        return (last + timedelta(days=1)).strftime("%Y%m%d")
    return "20250707"

# =============================================================================
# STEP 3: READ NEW JSON FILES — ONE PER DAY PER FACILITY
# =============================================================================

all_dfs     = []
total_files = 0

for local_path, spark_path, facility_name in FACILITY_SHORTCUTS:
    if not os.path.exists(local_path):
        print(f"Skipping missing shortcut: {local_path}")
        continue

    process_from  = get_process_from(facility_name)
    files_by_date = {}

    for fname in os.listdir(local_path):
        if not fname.endswith(".json"):
            continue
        try:
            file_date = fname.split("_")[1]  # YYYYMMDD
            if file_date >= process_from:
                # Keep only the latest file per day (max alphabetically = latest timestamp)
                if file_date not in files_by_date or fname > files_by_date[file_date]:
                    files_by_date[file_date] = fname
        except Exception:
            pass

    facility_files = [f"{spark_path}/{fname}" for fname in files_by_date.values()]

    if not facility_files:
        print(f"{facility_name}: No new files since {process_from} — skipping")
        continue

    print(f"{facility_name}: {len(facility_files)} new files (from {process_from})")
    total_files += len(facility_files)

    df = spark.read.option("multiline", "true").json(facility_files)

    df = df.withColumn("source_file", input_file_name())
    df = df.withColumn("file_date",
        to_date(regexp_extract(col("source_file"), r"census_(\d{8})", 1), "yyyyMMdd"))

    if "census_date" in df.columns:
        df = df.withColumn("census_date_final",
            coalesce(to_date(col("census_date")), col("file_date")))
    else:
        df = df.withColumn("census_date_final", col("file_date"))

    df_patients = df.select(
        col("census_date_final").alias("census_date"),
        col("extraction_timestamp"),
        col("record_count"),
        lit(facility_name).alias("facility"),
        explode(col("data.patients")).alias("patient")
    )

    df_flat = df_patients.select(
        col("census_date"),
        col("extraction_timestamp"),
        col("facility"),
        col("patient.casefile_id"),
        col("patient.mr_number"),
        col("patient.first_name"),
        col("patient.last_name"),
        col("patient.dob"),
        col("patient.gender"),
        col("patient.race"),
        col("patient.ethnicity"),
        to_date(col("patient.admission_date")).alias("admission_date"),
        to_date(col("patient.discharge_date")).alias("discharge_date"),
        regexp_replace(
            regexp_replace(
                regexp_replace(trim(col("patient.discharge_type")), "\\t", ""),
            "\u2019", "'"),
        "\u2013", "-").alias("discharge_type"),
        col("patient.discharge_type_code"),
        col("patient.discharge_or_transition_name"),
        to_date(col("patient.anticipated_discharge_date")).alias("anticipated_discharge_date"),
        col("patient.level_of_care"),
        col("patient.next_level_of_care"),
        trim(col("patient.program")).alias("program"),
        col("patient.location_id"),
        col("patient.location_name"),
        col("patient.building_name"),
        col("patient.room_name"),
        col("patient.bed_name"),
        trim(col("patient.insurance_company")).alias("insurance_company"),
        col("patient.payment_method"),
        col("patient.payment_method_category"),
        col("patient.referrer_name"),
        col("patient.first_contact_name"),
        col("patient.pre_admission_status"),
        col("patient.address_city"),
        col("patient.state"),
        col("patient.address_zip"),
        col("patient.preferred_language"),
        col("patient.diagnosis_codes"),
        col("patient.date_of_death"),
        col("patient.cause_of_death"),
        col("patient.created_at"),
        col("patient.last_updated_at")
    )

    all_dfs.append(df_flat)

if all_dfs:
    df_combined = reduce(lambda a, b: a.unionByName(b, allowMissingColumns=True), all_dfs)
    new_count = df_combined.count()
    print(f"\nTotal new records across all facilities: {new_count}")
    if incremental:
        df_combined.write.mode("append").format("delta").saveAsTable("census_raw")
    else:
        df_combined.write.mode("overwrite").option("overwriteSchema", "true").format("delta").saveAsTable("census_raw")
    print(f"✅ census_raw updated ({new_count} new records)")
else:
    print("No new files. Skipping census_raw append, rebuilding census_clean from existing data.")

# =============================================================================
# STEP 4: BUILD MAPPING SPARK DATAFRAMES
# =============================================================================

# Program mapping — 3-column key
program_schema = StructType([
    StructField("program_key",   StringType(), True),
    StructField("facility_key",  StringType(), True),
    StructField("location_key",  StringType(), True),
    StructField("program_group", StringType(), True),
])
program_df = spark.createDataFrame(program_mapping_rows, schema=program_schema)

# Insurance & discharge — single key
insurance_df = spark.createDataFrame(
    [(k, v) for k, v in insurance_mapping.items()],
    ["insurance_company", "payor_group"]
)
discharge_df = spark.createDataFrame(
    [(k, v) for k, v in discharge_mapping.items()],
    ["discharge_type", "discharge_group"]
)

# =============================================================================
# STEP 5: BUILD CENSUS_CLEAN
# =============================================================================

census_raw = spark.table("census_raw")

# Casefile IDs to keep despite matching the test-name filter.
# Use this for real patients whose KIPU MR# was set to a placeholder like "TEST"
# before the actual MR# was assigned.
# Format: {casefile_id: "facility — name — reason (date)"}
KEEP_DESPITE_TEST_FILTER = {
    "683:1e1f83ad-ad1d-4e41-b3b8-b9595ad42342": "RNAL — Joshua Fowler — admitted 2026-05-11, PHP at Central Alabama; intake used 'TEST2026-10' as placeholder MR# (logged 2026-05-26)",
}

# Remove test patients, with explicit allowlist for known false positives
keep_allowlist = (
    col("casefile_id").isin(list(KEEP_DESPITE_TEST_FILTER.keys()))
    if KEEP_DESPITE_TEST_FILTER else lit(False)
)

census_clean = census_raw.filter(
    keep_allowlist |
    (~(lower(col("first_name")).contains("test")) &
     ~(lower(col("last_name")).contains("test"))  &
     ~(lower(col("mr_number")).contains("test")))
)

# Deduplicate — one record per patient per census date per facility
census_clean = census_clean.dropDuplicates(["casefile_id", "census_date", "facility"])

# ── PROGRAM GROUP — 3-column join with fallback ───────────────────────────────
census_clean = census_clean \
    .withColumn("program_key",  coalesce(trim(col("program")),       lit("NULL"))) \
    .withColumn("facility_key", coalesce(trim(col("facility")),      lit(""))) \
    .withColumn("location_key", coalesce(trim(col("location_name")), lit("")))

# Pass 1: exact match on all 3 columns
exact_df = program_df.filter(
    (col("facility_key") != "") | (col("location_key") != "")
).select(
    col("program_key"),
    col("facility_key"),
    col("location_key"),
    col("program_group").alias("program_group_exact")
)
census_clean = census_clean.join(exact_df, on=["program_key", "facility_key", "location_key"], how="left")

# Pass 2: fallback — program name only (rows where facility and location are blank)
fallback_df = program_df.filter(
    (col("facility_key") == "") & (col("location_key") == "")
).select(
    col("program_key"),
    col("program_group").alias("program_group_fallback")
).dropDuplicates(["program_key"])
census_clean = census_clean.join(fallback_df, on="program_key", how="left")

# Use exact match first, fall back to program-only
census_clean = census_clean.withColumn("program_group",
    coalesce(col("program_group_exact"), col("program_group_fallback"))
).drop("program_group_exact", "program_group_fallback",
       "program_key", "facility_key", "location_key")

# ── INSURANCE & DISCHARGE ─────────────────────────────────────────────────────
census_clean = census_clean \
    .join(insurance_df, on="insurance_company", how="left") \
    .join(discharge_df, on="discharge_type",    how="left")

# Flag unmapped
census_clean = census_clean \
    .withColumn("program_group",
        when(col("program_group").isNull(), lit("⚠️ UNMAPPED"))
        .otherwise(col("program_group"))
    ) \
    .withColumn("payor_group",
        when(col("insurance_company").isNull() | (trim(col("insurance_company")) == ""), lit("Other"))
        .when(col("payor_group").isNull(), lit("⚠️ UNMAPPED"))
        .otherwise(col("payor_group"))
    ) \
    .withColumn("discharge_group",
    when(col("discharge_type").isNull() | (trim(col("discharge_type")) == ""), lit(None))
    .when(col("discharge_group").isNotNull(), col("discharge_group"))
    .when(col("discharge_date").isNotNull(), lit("⚠️ UNMAPPED"))
    .otherwise(lit(None)))
# Time dimensions
census_clean = census_clean \
    .withColumn("day_of_week",  date_format(col("census_date"), "EEEE")) \
    .withColumn("month",        date_format(col("census_date"), "yyyy-MM")) \
    .withColumn("week_ending",  next_day(col("census_date"), "SUN")) \
    .withColumn("year",         year(col("census_date")))

# Budget category — Longbranch only, all others = Unbudgeted
census_clean = census_clean.withColumn("budget_category",
    when(col("facility") != "Longbranch", lit("Unbudgeted"))
    .when(
        (col("location_name") == "Longbranch Recovery Center") &
        (col("program_group") == "Residential") &
        (col("payor_group") == "VA"),
        lit("VA Abita")
    ).when(
        (col("location_name") == "Longbranch Recovery Center") &
        (col("program_group") == "Residential") &
        (col("payor_group") != "VA"),
        lit("Other Abita")
    ).when(
        col("program_group") == "NORA",
        lit("NORA")
    ).when(
        (col("location_name").contains("Covington") | col("location_name").contains("Outpatient")) &
        (col("program_group") == "IOP") &
        (col("payor_group") == "VA"),
        lit("NSIOP VA")
    ).when(
        (col("location_name").contains("Covington") | col("location_name").contains("Outpatient")) &
        (col("program_group") == "IOP") &
        (col("payor_group") != "VA"),
        lit("NSIOP Other")
    ).when(
        col("location_name") == "Eating Disorder Treatment Center",
        lit("EDTC")
    ).otherwise(lit("Unbudgeted"))
)

census_clean.write.mode("overwrite").option("overwriteSchema", "true").format("delta").saveAsTable("census_clean")
total_clean = census_clean.count()
print(f"✅ census_clean updated ({total_clean} total records)")

# Facility breakdown
print("\nRecords per facility:")
census_clean.groupBy("facility").agg(
    count("*").alias("patient_days"),
    countDistinct("casefile_id").alias("unique_patients"),
    min("census_date").alias("from_date"),
    max("census_date").alias("to_date")
).orderBy("facility").show(20, truncate=False)

# =============================================================================
# STEP 6: BUDGET TABLE
# =============================================================================

budget_schema = StructType([
    StructField("month",           StringType(), True),
    StructField("budget_category", StringType(), True),
    StructField("budget_adc",      DoubleType(), True),
])
budget_df = spark.createDataFrame(budget_rows, schema=budget_schema)
budget_totals = budget_df.groupBy("month").agg(sum("budget_adc").alias("budget_adc_total"))
budget_df = budget_df.join(budget_totals, on="month", how="left")
budget_df.write.mode("overwrite").option("overwriteSchema", "true").format("delta").saveAsTable("budget")
print(f"✅ budget updated ({budget_df.count()} rows)")

# =============================================================================
# STEP 7: HELPER — WRITE UNMAPPED VALUES BACK TO EXCEL
# =============================================================================

def _write_unmapped_to_excel(sheet_name, new_rows_df):
    """Append unmapped values to the Lakehouse Excel mapping file."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        file_path = "/lakehouse/default/Files/Mappings/Longbranch_KIPU_Mapping.xlsx"
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        last_row = ws.max_row

        # program_mapping uses a 3-column composite key (Program + Facility + Location).
        # insurance_mapping and discharge_mapping use Program alone (column A).
        key_cols = 3 if sheet_name == "program_mapping" else 1

        existing_keys = set()
        for row in ws.iter_rows(min_row=3, max_col=key_cols, values_only=True):
            if row[0]:
                key = tuple((str(c).strip() if c else "") for c in row)
                existing_keys.add(key)

        THIN   = Side(style="thin", color="DDDDDD")
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ALERT  = PatternFill("solid", start_color="FFF3CD")
        AFONT  = Font(name="Arial", size=10, color="856404", bold=True)
        NFONT  = Font(name="Arial", size=10, color="856404")

        added = 0
        for _, row in new_rows_df.iterrows():
            new_key = tuple(
                (str(row.iloc[i]).strip() if pd.notna(row.iloc[i]) else "")
                for i in range(key_cols)
            )
            if new_key in existing_keys:
                continue
            last_row += 1
            for col_idx, val in enumerate(row, 1):
                c = ws.cell(last_row, col_idx)
                c.value = str(val) if pd.notna(val) and val != "" else ""
                c.fill = ALERT
                c.font = AFONT if col_idx == len(row) else NFONT
                c.alignment = Alignment(horizontal="left", vertical="center")
                c.border = BORDER
            added += 1

        if added > 0:
            wb.save(file_path)
            print(f"  → {added} new unmapped values written to {sheet_name} in Excel")
        else:
            print(f"  → No new unmapped values for {sheet_name}")

    except Exception as e:
        print(f"  ⚠️ Could not write to Excel: {str(e)}")

# =============================================================================
# STEP 8: VALIDATION
# =============================================================================

unmapped_programs  = census_clean.filter(col("program_group")  == "⚠️ UNMAPPED")
unmapped_insurance = census_clean.filter(col("payor_group")    == "⚠️ UNMAPPED")
unmapped_discharge = census_clean.filter(col("discharge_group") == "⚠️ UNMAPPED")


total_issues = unmapped_programs.count() + unmapped_insurance.count() + unmapped_discharge.count()
status = "⚠️ ISSUES FOUND — check Excel mapping file" if total_issues > 0 else "✅ ALL CLEAN"

print(f"\n=== VALIDATION REPORT ===")

if unmapped_programs.count() > 0:
    print(f"⚠️  UNMAPPED PROGRAMS: {unmapped_programs.count()} patient-days")
    up = unmapped_programs.groupBy("program", "facility", "location_name").agg(
        count("*").alias("patient_days"),
        countDistinct("casefile_id").alias("unique_patients"),
        max("census_date").alias("last_seen")
    ).orderBy(desc("patient_days"))
    up.show(30, truncate=False)
    up_pd = up.toPandas()
    up_pd.columns = ["Program Name (from Kipu)", "Facility", "Location", "Patient Days", "Unique Patients", "Last Seen"]
    up_pd["Program Group"] = ""
    up_pd["Status"] = "⚠️ Unmapped — add group"
    _write_unmapped_to_excel("program_mapping", up_pd[["Program Name (from Kipu)", "Facility", "Location", "Program Group", "Status"]])
else:
    print("✅ All programs mapped")

if unmapped_insurance.count() > 0:
    print(f"⚠️  UNMAPPED INSURANCE: {unmapped_insurance.count()} patient-days")
    ui = unmapped_insurance.groupBy("insurance_company", "facility").agg(
        count("*").alias("patient_days"),
        countDistinct("casefile_id").alias("unique_patients"),
        max("census_date").alias("last_seen")
    ).orderBy(desc("patient_days"))
    ui.show(30, truncate=False)
    ui_pd = ui.toPandas()
    ui_pd.columns = ["Insurance Company (from Kipu)", "Facility", "Patient Days", "Unique Patients", "Last Seen"]
    ui_pd["Payor Group"] = ""
    ui_pd["Status"] = "⚠️ Unmapped — add group"
    _write_unmapped_to_excel("insurance_mapping", ui_pd[["Insurance Company (from Kipu)", "Payor Group", "Status"]])
else:
    print("✅ All insurance companies mapped")

if unmapped_discharge.count() > 0:
    print(f"⚠️  UNMAPPED DISCHARGE TYPES: {unmapped_discharge.count()} patient-days")
    ud = unmapped_discharge.groupBy("discharge_type", "facility").agg(
        count("*").alias("patient_days"),
        countDistinct("casefile_id").alias("unique_patients"),
        max("census_date").alias("last_seen")
    ).orderBy(desc("patient_days"))
    ud.show(30, truncate=False)
    ud_pd = ud.toPandas()
    ud_pd.columns = ["Discharge Type (from Kipu)", "Facility", "Patient Days", "Unique Patients", "Last Seen"]
    ud_pd["Discharge Group"] = ""
    ud_pd["Status"] = "⚠️ Unmapped — add group"
    _write_unmapped_to_excel("discharge_mapping", ud_pd[["Discharge Type (from Kipu)", "Discharge Group", "Status"]])
else:
    print("✅ All discharge types mapped")

validation_df = spark.createDataFrame([
    (
        unmapped_programs.count(),
        unmapped_insurance.count(),
        unmapped_discharge.count(),
        total_clean,
        status,
        str(dt.now(timezone.utc))
    )
], ["unmapped_program_days", "unmapped_insurance_days", "unmapped_discharge_days",
    "total_patient_days", "data_status", "last_checked"])

validation_df.write.mode("overwrite").option("overwriteSchema", "true").format("delta").saveAsTable("data_validation")
print(f"✅ data_validation updated")

# =============================================================================
# FINAL SUMMARY
# =============================================================================

census_final = spark.table("census_clean")
date_range   = census_final.select(min("census_date"), max("census_date")).first()

print("\n" + "="*60)
print("PIPELINE COMPLETE")
print("="*60)
print(f"Total patient-days (census_clean) : {census_final.count()}")
print(f"Date range                        : {date_range[0]} to {date_range[1]}")
print(f"Unique patients                   : {census_final.select('casefile_id').distinct().count()}")
print(f"Facilities                        : {census_final.select('facility').distinct().count()}")
print(f"Data status                       : {status}")
print(f"Run completed                     : {dt.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
print("\nTables updated: census_raw, census_clean, budget, data_validation")

# METADATA ********************

# META {
# META   "language": "python",
# META   "language_group": "synapse_pyspark"
# META }
